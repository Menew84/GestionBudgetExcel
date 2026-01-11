
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def creer_fichier_budget(nom_fichier="Budget_Personnel.xlsx"):
    annee = "2026"
    mois_fr = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    # === PARAMÈTRES CLÉS — étendue de saisie agrandie ===
    ENTRY_START_ROW = 6
    ENTRY_END_ROW   = 106  # <- anciennement 55
    HIDDEN_ROW      = ENTRY_END_ROW + 4  # ligne technique pour totaux/solde (hors zone de saisie)

    # Formats
    monnaie_format = '#,##0.00 €;[Red]-#,##0.00 €'   # Revenus (positifs, rouge si négatif)
    depense_format = '"-"#,##0.00 €'                 # Dépenses stockées en positif, affichées en "-"

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
    fill_red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # ----------------------------------------------------------------
    # 1. PAGE DE GARDE
    # ----------------------------------------------------------------
    page_garde = wb.create_sheet("Page de Garde")

    page_garde.merge_cells("A1:D1")
    page_garde["A1"] = f"BUDGET PERSONNEL {annee} - Récapitulatif Annuel"
    page_garde["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    page_garde["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    page_garde["A1"].alignment = Alignment(horizontal="center")

    # En-têtes (ligne 3)
    headers = [("Mois", "A3"), ("Total Revenus", "B3"), ("Total Dépenses", "C3"), ("Solde", "D3")]
    for text, cell_ref in headers:
        cell = page_garde[cell_ref]
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Lignes 4..15 = 12 mois
    start_row = 4
    for i, mois in enumerate(mois_fr, start=start_row):
        row = i
        page_garde.cell(row=row, column=1, value=f"{mois} {annee}")
        # Colonnes B,C,D pointent vers la ligne HIDDEN_ROW de chaque mois
        page_garde.cell(row=row, column=2, value=f"='{mois} {annee}'!C{HIDDEN_ROW}")  # Total Revenus
        page_garde.cell(row=row, column=3, value=f"='{mois} {annee}'!D{HIDDEN_ROW}")  # Total Dépenses
        page_garde.cell(row=row, column=4, value=f"='{mois} {annee}'!E{HIDDEN_ROW}")  # Solde Mensuel

        for c in range(1, 5):
            cell_ = page_garde.cell(row=row, column=c)
            if c in (2, 3, 4):
                cell_.number_format = monnaie_format
            cell_.border = thin_border

    # Ligne 16 = "TOTAL Annuel"
    total_row = start_row + len(mois_fr)
    page_garde.cell(row=total_row, column=1, value="TOTAL Annuel").font = Font(bold=True)
    page_garde.cell(row=total_row, column=2, value=f"=SUM(B{start_row}:B{total_row-1})")
    page_garde.cell(row=total_row, column=3, value=f"=SUM(C{start_row}:C{total_row-1})")
    # Pour la colonne D, on récupère le solde du dernier mois (ligne 15)
    page_garde.cell(row=total_row, column=4, value=f"=D15")

    for c in range(1, 5):
        cell_ = page_garde.cell(row=total_row, column=c)
        if c in (2, 3, 4):
            cell_.number_format = monnaie_format
        cell_.border = thin_border

    # Largeurs
    page_garde.column_dimensions["A"].width = 30
    page_garde.column_dimensions["B"].width = 20
    page_garde.column_dimensions["C"].width = 20
    page_garde.column_dimensions["D"].width = 20

    # Graphique
    chart = BarChart()
    chart.title = "Comparatif Revenus / Dépenses"
    chart.x_axis.title = "Mois"
    chart.y_axis.title = "Montant (€)"

    # On ignore la ligne "TOTAL Annuel" => max_row=15
    data = Reference(page_garde, min_col=2, max_col=3, min_row=3, max_row=15)
    cats = Reference(page_garde, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.type = "col"
    chart.grouping = "clustered"
    chart.width = 30
    chart.height = 15
    chart.legend.position = "b"

    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showCatName = True
        s.dLbls.showSerName = False
        s.dLbls.number_format = monnaie_format
        s.dLbls.separator = "\\n"

    chart.x_axis.label_rotation = 45
    chart.y_axis.number_format = monnaie_format
    chart.gapWidth = 150
    page_garde.add_chart(chart, "F3")

    # ----------------------------------------------------------------
    # 2. FEUILLES MENSUELLES
    # ----------------------------------------------------------------
    for i, mois in enumerate(mois_fr):
        nom_feuille = f"{mois} {annee}"
        ws = wb.create_sheet(nom_feuille)

        # Titre
        ws.merge_cells("A1:E1")
        ws["A1"] = f"{mois.upper()} {annee}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Largeurs
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 9
        ws.column_dimensions["E"].width = 2

        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 25
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 9

        # Solde Initial (ligne 2)
        ws["A2"] = "Solde Initial"
        ws["A2"].font = Font(bold=True)
        ws["A2"].border = thin_border

        if i == 0:
            ws["B2"] = 0
        else:
            ws["B2"] = f"='{mois_fr[i-1]} {annee}'!C4"
        ws["B2"].number_format = monnaie_format
        ws["B2"].border = thin_border

        # Ligne 3 : Totaux
        ws.merge_cells("A3:B3")
        ws["A3"] = "Total Revenus"
        ws["A3"].font = Font(bold=True)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws["A3"].fill = fill_green
        ws["A3"].border = thin_border

        ws["C3"] = f"=SUM(C{ENTRY_START_ROW}:C{ENTRY_END_ROW})"
        ws["C3"].font = Font(bold=True)
        ws["C3"].number_format = monnaie_format
        ws["C3"].border = thin_border

        ws.merge_cells("F3:G3")
        ws["F3"] = "Total Dépenses"
        ws["F3"].font = Font(bold=True)
        ws["F3"].alignment = Alignment(horizontal="center")
        ws["F3"].fill = fill_red
        for col_ in range(6, 8):  # F..G
            ws.cell(row=3, column=col_).border = thin_border

        ws["H3"] = f"=SUM(H{ENTRY_START_ROW}:H{ENTRY_END_ROW})"
        ws["H3"].font = Font(bold=True)
        ws["H3"].number_format = depense_format
        ws["H3"].border = thin_border

        # Ligne 4 : Solde Mensuel
        ws["B4"] = "Solde Mensuel"
        ws["B4"].font = Font(bold=True)
        ws["B4"].alignment = Alignment(horizontal="center")
        ws["B4"].border = thin_border

        ws["C4"] = "=B2 + C3 - H3"
        ws["C4"].font = Font(bold=True)
        ws["C4"].number_format = monnaie_format
        ws["C4"].border = thin_border

        # Ligne 5 : En-têtes
        # Revenus
        ws["A5"] = "Date"
        ws["B5"] = "Catégorie/Desc."
        ws["C5"] = "Montant"
        ws["D5"] = "Payé ?"
        for col_ in range(1, 5):
            cell_ = ws.cell(row=5, column=col_)
            cell_.font = Font(bold=True)
            cell_.alignment = Alignment(horizontal="center")
            cell_.fill = fill_green
            cell_.border = thin_border

        # Dépenses
        ws["F5"] = "Date"
        ws["G5"] = "Catégorie/Desc."
        ws["H5"] = "Montant"
        ws["I5"] = "Payé ?"
        for col_ in range(6, 10):
            cell_ = ws.cell(row=5, column=col_)
            cell_.font = Font(bold=True)
            cell_.alignment = Alignment(horizontal="center")
            cell_.fill = fill_red
            cell_.border = thin_border

        # Zone de saisie 6..106
        for row_ in range(ENTRY_START_ROW, ENTRY_END_ROW + 1):
            # Revenus
            ws.cell(row=row_, column=1).number_format = "dd/mm/yyyy"
            ws.cell(row=row_, column=1).border = thin_border
            ws.cell(row=row_, column=2).border = thin_border
            rev_mont = ws.cell(row=row_, column=3)
            rev_mont.number_format = monnaie_format
            rev_mont.border = thin_border
            ws.cell(row=row_, column=4).border = thin_border

            # Dépenses
            ws.cell(row=row_, column=6).number_format = "dd/mm/yyyy"
            ws.cell(row=row_, column=6).border = thin_border
            ws.cell(row=row_, column=7).border = thin_border
            dep_mont = ws.cell(row=row_, column=8)
            dep_mont.number_format = depense_format
            dep_mont.border = thin_border
            dep_mont.font = Font(color="FF0000")  # texte rouge
            ws.cell(row=row_, column=9).border = thin_border

        # Cellules techniques (ligne HIDDEN_ROW)
        ws[f"C{HIDDEN_ROW}"] = "=C3"  # Total Revenus
        ws[f"D{HIDDEN_ROW}"] = "=H3"  # Total Dépenses
        ws[f"E{HIDDEN_ROW}"] = "=C4"  # Solde
        ws.row_dimensions[HIDDEN_ROW].hidden = True

    # Forcer le recalcul
    wb.excel_calc_id = 0
    for ws in wb.worksheets:
        ws.sheet_properties.forceFormulaRecalculation = True

    wb.save(nom_fichier)
    print(f"Fichier '{nom_fichier}' créé avec succès !")

if __name__ == "__main__":
    creer_fichier_budget("Budget_Personnel.xlsx")
